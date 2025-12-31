"""
Excel report generator for the activity reporting system.

Creates formatted Excel reports with multiple sheets, charts, and styling.
"""

import pandas as pd
from pathlib import Path
from typing import Dict, Any, Optional, List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from ..utils.logger import setup_logger

logger = setup_logger(__name__)


class ExcelReportGenerator:
    """
    Generate formatted Excel reports.

    Creates multi-sheet Excel workbooks with:
    - Formatted tables
    - Summary statistics
    - Charts and visualizations
    - Custom styling

    Example:
        >>> generator = ExcelReportGenerator()
        >>> generator.generate(summaries, 'reports/weekly_report.xlsx')
    """

    def __init__(self):
        """Initialize the Excel report generator."""
        # Define color scheme
        self.colors = {
            'header': 'FF4472C4',
            'subheader': 'FFD9E1F2',
            'highlight': 'FFFFFF00',
            'positive': 'FFC6EFCE',
            'negative': 'FFFFC7CE'
        }

    def generate(
        self,
        summaries: Dict[str, pd.DataFrame],
        output_path: str,
        config: Optional[Dict[str, Any]] = None
    ) -> str:
        """
        Generate Excel report from summary data.

        Args:
            summaries: Dictionary of summary DataFrames
            output_path: Path for output Excel file
            config: Report configuration options

        Returns:
            Path to generated Excel file

        Configuration options:
            - title: Report title (default: 'Weekly Activity Report')
            - include_charts: Add charts to report (default: True)
            - freeze_panes: Freeze header rows (default: True)
        """
        if config is None:
            config = {}

        logger.info(f"Generating Excel report: {output_path}")

        # Create output directory if needed
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Add summary sheet
        self._add_summary_sheet(wb, summaries, config)

        # Add detailed sheets for each summary
        for sheet_name, df in summaries.items():
            if isinstance(df, pd.DataFrame) and not df.empty:
                self._add_data_sheet(wb, sheet_name, df, config)

        # Add charts if requested
        if config.get('include_charts', True):
            self._add_charts(wb, summaries)

        # Save workbook
        wb.save(output_path)

        logger.info(f"Excel report generated successfully: {output_path}")
        return output_path

    def _add_summary_sheet(
        self,
        wb: Workbook,
        summaries: Dict[str, pd.DataFrame],
        config: Dict[str, Any]
    ) -> None:
        """Add overview summary sheet."""
        ws = wb.create_sheet('Summary', 0)

        # Add title
        title = config.get('title', 'Weekly Activity Report')
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.colors['header'],
                                     end_color=self.colors['header'],
                                     fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')

        # Add generation date
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)

        row = 4

        # Add summary information
        ws[f'A{row}'] = "Report Contents:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        for sheet_name, df in summaries.items():
            if isinstance(df, pd.DataFrame):
                ws[f'B{row}'] = f"• {sheet_name.replace('_', ' ').title()}"
                ws[f'C{row}'] = f"{len(df)} rows"
                row += 1

        # Add key metrics if available
        if 'weekly_totals' in summaries and not summaries['weekly_totals'].empty:
            row += 2
            ws[f'A{row}'] = "Latest Week Summary:"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

            latest = summaries['weekly_totals'].iloc[-1]
            for col, value in latest.items():
                if col not in ['week_start', 'week_number', 'year']:
                    ws[f'B{row}'] = col.replace('_', ' ').title()
                    ws[f'C{row}'] = value
                    row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def _add_data_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        df: pd.DataFrame,
        config: Dict[str, Any]
    ) -> None:
        """Add a data sheet with formatted table."""
        # Clean sheet name (Excel has 31 char limit and special char restrictions)
        clean_name = sheet_name.replace('_', ' ').title()[:31]
        ws = wb.create_sheet(clean_name)

        # Write DataFrame to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Style header row
                if r_idx == 1:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color=self.colors['header'],
                                           end_color=self.colors['header'],
                                           fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')

                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border

                # Format numbers
                if r_idx > 1 and isinstance(value, (int, float)):
                    if abs(value) < 1 and value != 0:
                        cell.number_format = '0.00'
                    elif abs(value) >= 1000:
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '0.00'

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row if requested
        if config.get('freeze_panes', True):
            ws.freeze_panes = 'A2'

    def _add_charts(
        self,
        wb: Workbook,
        summaries: Dict[str, pd.DataFrame]
    ) -> None:
        """Add charts to visualize data."""
        # Add weekly trend chart if data available
        if 'weekly_totals' in summaries and not summaries['weekly_totals'].empty:
            try:
                self._add_weekly_trend_chart(wb, summaries['weekly_totals'])
            except Exception as e:
                logger.warning(f"Failed to add weekly trend chart: {str(e)}")

    def _add_weekly_trend_chart(
        self,
        wb: Workbook,
        weekly_df: pd.DataFrame
    ) -> None:
        """Add line chart showing weekly trends."""
        # Find the weekly totals sheet
        sheet_name = 'Weekly Totals'
        if sheet_name not in wb.sheetnames:
            return

        ws = wb[sheet_name]

        # Get numeric columns (excluding week identifiers)
        numeric_cols = []
        for idx, col in enumerate(weekly_df.columns, 1):
            if col not in ['week_start', 'week_number', 'year'] and \
               pd.api.types.is_numeric_dtype(weekly_df[col]):
                numeric_cols.append((idx, col))

        if not numeric_cols:
            return

        # Create line chart for first numeric column
        chart = LineChart()
        chart.title = "Weekly Trends"
        chart.style = 10
        chart.y_axis.title = 'Value'
        chart.x_axis.title = 'Week'

        # Add data series
        col_idx, col_name = numeric_cols[0]
        data = Reference(ws, min_col=col_idx, min_row=1, max_row=len(weekly_df) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(weekly_df) + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Add chart to sheet
        ws.add_chart(chart, f'A{len(weekly_df) + 5}')

        logger.info("Added weekly trend chart to report")
